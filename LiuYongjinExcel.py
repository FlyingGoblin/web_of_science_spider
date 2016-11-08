# -*- coding: utf-8 -*
import openpyxl
from re import sub

from PaperExcel import PaperExcel
from My_enum import enum

__author__ = 'flyinggoblin'

TABLE_TITLE = enum('zero',
                   'TITLE',
                   'JOURNAL',
                   'AUTHORS',
                   'YEAR',
                   'PAPER_COUNTER',
                   'IS_FIRST_AUTHOR',
                   'IS_SECOND_AUTHOR',
                   'IS_COR_AUTHOR',
                   'IDS_NUMBER',
                   'CITE_NUMBER',
                   'OTHER_CITE_NUMBER',
                   'CITE_TITLE',
                   'CITE_JOURNAL',
                   'CITE_AUTHOR',
                   'CITE_YEAR',
                   'CITE_COUNTER',
                   'IS_OTHER_CITE')

TABLE_TITLE_CN = {
    TABLE_TITLE.TITLE: u'题目',
    TABLE_TITLE.JOURNAL: u'收录',
    TABLE_TITLE.AUTHORS: u'作者',
    TABLE_TITLE.YEAR: u'时间(年)',
    TABLE_TITLE.PAPER_COUNTER: u'PAPER_COUNTER',
    TABLE_TITLE.IS_FIRST_AUTHOR: u'一作',
    TABLE_TITLE.IS_SECOND_AUTHOR: u'二作',
    TABLE_TITLE.IS_COR_AUTHOR: u'通讯作者',
    TABLE_TITLE.IDS_NUMBER: u'检索号',
    TABLE_TITLE.CITE_NUMBER: u'引用',
    TABLE_TITLE.OTHER_CITE_NUMBER: u'他引',
    TABLE_TITLE.CITE_TITLE: u'引用文章',
    TABLE_TITLE.CITE_JOURNAL: u'引用文章期刊',
    TABLE_TITLE.CITE_AUTHOR: u'引用作者',
    TABLE_TITLE.CITE_YEAR: u'引用时间(年)',
    TABLE_TITLE.CITE_COUNTER: u'CITE_COUNTER',
    TABLE_TITLE.IS_OTHER_CITE: u'是否他引'
}

FIRST_AUTHOR = TABLE_TITLE.IS_FIRST_AUTHOR
SECOND_AUTHOR = TABLE_TITLE.IS_SECOND_AUTHOR
COR_AUTHOR = TABLE_TITLE.IS_COR_AUTHOR


class LiuYongjinExcel(PaperExcel):
    def __init__(self, excel_path):
        self.__excel_path = excel_path
        self.__excel = openpyxl.Workbook()
        self.__table = self.__excel.active
        self.__write_table_title()

    def __write_table_title(self):
        titles = (TABLE_TITLE_CN[col] for col in TABLE_TITLE_CN)
        self.__table.append(titles)
        self.__excel.save(self.__excel_path)

    def write_paper_cite(self, paper, cites, authors_tag):
        row_start_paper = self.__table.max_row + 1
        # 论文本身
        self.__table.cell(row=row_start_paper, column=TABLE_TITLE.TITLE).value = paper.title
        self.__table.cell(row=row_start_paper, column=TABLE_TITLE.JOURNAL).value = paper.journal
        row_author = row_start_paper
        for author in paper.authors:
            self.__table.cell(row=row_author, column=TABLE_TITLE.AUTHORS).value = author
            row_author += 1
        self.__table.cell(row=row_start_paper, column=TABLE_TITLE.YEAR).value = paper.year
        self.__table.cell(row=row_start_paper, column=TABLE_TITLE.IDS_NUMBER).value = paper.ids
        # 专享
        self.__table.cell(row=row_start_paper, column=TABLE_TITLE.PAPER_COUNTER).value = 1
        for author_tag in authors_tag:
            self.__table.cell(row=row_start_paper, column=author_tag).value = 1
        # 引用信息
        self.__table.cell(row=row_start_paper, column=TABLE_TITLE.CITE_NUMBER).value = len(cites)
        other_cite_count = 0
        row_start_cite = row_start_paper
        for cite in cites:
            self.__table.cell(row=row_start_cite, column=TABLE_TITLE.CITE_TITLE).value = cite.title
            self.__table.cell(row=row_start_cite, column=TABLE_TITLE.CITE_JOURNAL).value = cite.journal
            row_author = row_start_cite
            for author in cite.authors:
                self.__table.cell(row=row_author, column=TABLE_TITLE.CITE_AUTHOR).value = author
                row_author += 1
            self.__table.cell(row=row_start_cite, column=TABLE_TITLE.CITE_YEAR).value = cite.year
            if paper.is_self_cite(cite):
                self.__table.cell(row=row_start_cite, column=TABLE_TITLE.IS_OTHER_CITE).value = 0
            else:
                self.__table.cell(row=row_start_cite, column=TABLE_TITLE.IS_OTHER_CITE).value = 1
                other_cite_count += 1
            # 专享
            self.__table.cell(row=row_start_cite, column=TABLE_TITLE.CITE_COUNTER).value = 1
            row_start_cite = self.__table.max_row + 1
        self.__table.cell(row=row_start_paper, column=TABLE_TITLE.OTHER_CITE_NUMBER).value = other_cite_count
        self.__excel.save(self.__excel_path)

    @classmethod
    def read_title_list_form_per_result(cls, excel_path):
        per_excel = openpyxl.load_workbook(excel_path)
        per_table = per_excel['SCI']
        rows = per_table.max_row
        paper_list = []
        for row in range(1, rows):
            true_row = row + 1
            if per_table.cell(row=true_row, column=TABLE_TITLE.TITLE).value:
                title = sub(r'\s+', ' ', per_table.cell(row=true_row, column=TABLE_TITLE.TITLE).value)
                author_tags = []
                if per_table.cell(row=true_row, column=TABLE_TITLE.IS_FIRST_AUTHOR).value:
                    author_tags.append(FIRST_AUTHOR)
                if per_table.cell(row=true_row, column=TABLE_TITLE.IS_SECOND_AUTHOR).value:
                    author_tags.append(SECOND_AUTHOR)
                if per_table.cell(row=true_row, column=TABLE_TITLE.IS_COR_AUTHOR).value:
                    author_tags.append(COR_AUTHOR)
                paper_list.append([title, author_tags])
        return paper_list
