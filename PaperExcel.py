# -*- coding: utf-8 -*
import openpyxl
import xlrd

from Paper import Paper
from My_enum import enum

__author__ = 'flyinggoblin'

TABLE_TITLE = enum('zero',
                   'TITLE',
                   'JOURNAL',
                   'AUTHORS',
                   'YEAR',
                   'CITE_NUMBER',
                   'OTHER_CITE_NUMBER',
                   'CITE_TITLE',
                   'CITE_JOURNAL',
                   'CITE_AUTHOR',
                   'CITE_YEAR',
                   'IS_SCI',
                   'IS_OTHER_CITE')

TABLE_TITLE_CN = {
    TABLE_TITLE.TITLE: u'题目',
    TABLE_TITLE.JOURNAL: u'收录',
    TABLE_TITLE.AUTHORS: u'作者',
    TABLE_TITLE.YEAR: u'时间(年)',
    TABLE_TITLE.CITE_NUMBER: u'引用',
    TABLE_TITLE.OTHER_CITE_NUMBER: u'他引',
    TABLE_TITLE.CITE_TITLE: u'引用文章',
    TABLE_TITLE.CITE_JOURNAL: u'引用文章期刊',
    TABLE_TITLE.CITE_AUTHOR: u'引用作者',
    TABLE_TITLE.CITE_YEAR: u'引用时间(年)',
    TABLE_TITLE.IS_SCI: u'是否SCI引用',
    TABLE_TITLE.IS_OTHER_CITE: u'是否他引',
}


class PaperExcel(object):
    def __init__(self, excel_path):
        self.__excel_path = excel_path
        self.__excel = openpyxl.Workbook()
        self.__table = self.__excel.active
        self.__write_table_title()

    def __write_table_title(self):
        titles = (TABLE_TITLE_CN[col] for col in TABLE_TITLE_CN)
        self.__table.append(titles)
        self.__excel.save(self.__excel_path)

    def write_paper_cite(self, paper, cites):
        row_start_paper = self.__table.max_row + 1
        # 论文本身
        self.__table.cell(row=row_start_paper, column=TABLE_TITLE.TITLE).value = paper.title
        self.__table.cell(row=row_start_paper, column=TABLE_TITLE.JOURNAL).value = paper.journal
        row_author = row_start_paper
        for author in paper.authors:
            self.__table.cell(row=row_author, column=TABLE_TITLE.AUTHORS).value = author
            row_author += 1
        self.__table.cell(row=row_start_paper, column=TABLE_TITLE.YEAR).value = paper.year
        # 引用信息
        self.__table.cell(row=row_start_paper, column=TABLE_TITLE.CITE_NUMBER).value = len(cites)
        other_cite_count = 0
        row_start_cite = row_start_paper
        for cite in cites:
            self.__table.cell(row=row_start_cite, column=TABLE_TITLE.CITE_TITLE).value = cite.title
            self.__table.cell(row=row_start_cite, column=TABLE_TITLE.CITE_JOURNAL).value = cite.journal
            row_author = row_start_cite
            for author in cite.authors:
                self.__table.cell(row=row_author, column=TABLE_TITLE.CITE_AUTHOR).value = author
                row_author += 1
            self.__table.cell(row=row_start_cite, column=TABLE_TITLE.CITE_YEAR).value = cite.year
            self.__table.cell(row=row_start_cite, column=TABLE_TITLE.IS_SCI).value = 1
            if paper.is_self_cite(cite):
                self.__table.cell(row=row_start_cite, column=TABLE_TITLE.IS_OTHER_CITE).value = 0
            else:
                self.__table.cell(row=row_start_cite, column=TABLE_TITLE.IS_OTHER_CITE).value = 1
                other_cite_count += 1
            row_start_cite = self.__table.max_row + 1
        self.__table.cell(row=row_start_paper, column=TABLE_TITLE.OTHER_CITE_NUMBER).value = other_cite_count
        self.__excel.save(self.__excel_path)

    @classmethod
    def read_title_list(cls, excel_path):
        title_data = xlrd.open_workbook(excel_path)
        table = title_data.sheets()[0]
        return table.col_values(0)
